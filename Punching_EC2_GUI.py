# -*- coding: utf-8 -*-
from __future__ import annotations

from datetime import datetime
import math
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from Punching_EC2 import PuncoamentoEC2

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfbase.pdfmetrics import stringWidth
    from reportlab.pdfgen import canvas as pdf_canvas
    REPORTLAB_OK = True
except Exception:
    REPORTLAB_OK = False

APP_TITLE = "PunchingShearEC2"
BG = "#f5f7fb"
CARD = "#ffffff"
ACCENT = "#2563eb"
TEXT = "#0f172a"
MUTED = "#475569"
OK = "#15803d"
WARN = "#b45309"
FAIL = "#b91c1c"

EXEMPLOS = {
    "Pilar interior retangular": {
        "fck": "30", "fyk": "500", "fywk": "500", "d": "0.22", "asx": "12.57", "asy": "12.57",
        "sigma_cp": "0", "tipo": "interior", "forma": "retangular", "c1": "0.40", "c2": "0.40",
        "ved": "600", "medx": "0", "medy": "0", "is_sapata": False, "sigma_gd": "150",
        "has_abertura": False, "u1_inef": "0", "beta": "simplificado", "edge_interior": True,
        "corner_interior": True,
    },
    "Pilar de bordo com momento": {
        "fck": "30", "fyk": "500", "fywk": "500", "d": "0.22", "asx": "10.50", "asy": "12.57",
        "sigma_cp": "0", "tipo": "bordo", "forma": "retangular", "c1": "0.40", "c2": "0.30",
        "ved": "550", "medx": "60", "medy": "25", "is_sapata": False, "sigma_gd": "150",
        "has_abertura": False, "u1_inef": "0", "beta": "ec2", "edge_interior": True,
        "corner_interior": True,
    },
    "Pilar de canto": {
        "fck": "35", "fyk": "500", "fywk": "500", "d": "0.24", "asx": "14.14", "asy": "14.14",
        "sigma_cp": "0", "tipo": "canto", "forma": "retangular", "c1": "0.40", "c2": "0.40",
        "ved": "450", "medx": "35", "medy": "20", "is_sapata": False, "sigma_gd": "150",
        "has_abertura": False, "u1_inef": "0", "beta": "ec2", "edge_interior": True,
        "corner_interior": True,
    },
    "Pilar circular": {
        "fck": "30", "fyk": "500", "fywk": "500", "d": "0.22", "asx": "12.57", "asy": "12.57",
        "sigma_cp": "0", "tipo": "interior", "forma": "circular", "c1": "0.40", "c2": "0.40",
        "ved": "600", "medx": "0", "medy": "0", "is_sapata": False, "sigma_gd": "150",
        "has_abertura": False, "u1_inef": "0", "beta": "fib", "edge_interior": True,
        "corner_interior": True,
    },
}


class PuncoamentoApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1180x760")
        self.minsize(960, 640)
        self.configure(bg=BG)
        self.last_report = ""
        self.last_verif = None
        self.drag_mode = None
        self._apply_theme()
        self._build_variables()
        self._build_ui()
        self._apply_visibility_rules()
        self._update_rho_label()
        self._draw_scheme()

    def _apply_theme(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("TFrame", background=BG)
        style.configure("Card.TFrame", background=CARD)
        style.configure("TLabelframe", background=CARD, foreground=TEXT)
        style.configure("TLabelframe.Label", background=CARD, foreground=TEXT, font=("Segoe UI", 10, "bold"))
        style.configure("TLabel", background=BG, foreground=TEXT, font=("Segoe UI", 10))
        style.configure("Card.TLabel", background=CARD, foreground=TEXT)
        style.configure("Muted.TLabel", background=CARD, foreground=MUTED, font=("Segoe UI", 9))
        style.configure("Header.TLabel", background=BG, foreground=TEXT, font=("Segoe UI", 15, "bold"))
        style.configure("TButton", padding=(10, 7), font=("Segoe UI", 10))
        style.configure("Accent.TButton", padding=(10, 7), font=("Segoe UI", 10, "bold"))
        style.map("Accent.TButton", background=[("!disabled", ACCENT)], foreground=[("!disabled", "white")])
        style.configure("TNotebook", background=BG, borderwidth=0)
        style.configure("TNotebook.Tab", padding=(12, 8), font=("Segoe UI", 10))
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
        style.configure("Treeview", rowheight=24, font=("Segoe UI", 10))
        style.configure("TCombobox", padding=4)
        style.configure("TEntry", padding=4)
        style.configure("Horizontal.TScale", background=CARD)

    def _build_variables(self):
        self.var_fck = tk.StringVar(value="30")
        self.var_fyk = tk.StringVar(value="500")
        self.var_fywk = tk.StringVar(value="500")
        self.var_d = tk.StringVar(value="0.22")
        self.var_asx = tk.StringVar(value="10.0")
        self.var_asy = tk.StringVar(value="10.0")
        self.var_sigma_cp = tk.StringVar(value="0")
        self.var_tipo_pilar = tk.StringVar(value="interior")
        self.var_forma_pilar = tk.StringVar(value="retangular")
        self.var_c1 = tk.StringVar(value="0.40")
        self.var_c2 = tk.StringVar(value="0.40")
        self.var_ved = tk.StringVar(value="600")
        self.var_medx = tk.StringVar(value="0")
        self.var_medy = tk.StringVar(value="0")
        self.var_is_sapata = tk.BooleanVar(value=False)
        self.var_sigma_gd = tk.StringVar(value="150")
        self.var_has_abertura = tk.BooleanVar(value=False)
        self.var_u1_inef = tk.StringVar(value="0")
        self.var_beta = tk.StringVar(value="simplificado")
        self.var_exemplo = tk.StringVar(value="Pilar interior retangular")
        self.var_edge_interior = tk.BooleanVar(value=True)
        self.var_corner_interior = tk.BooleanVar(value=True)
        self.var_rho_calc = tk.StringVar(value="ρl = -")
        self.var_status = tk.StringVar(value="Pronto.")
        self.var_resultado = tk.StringVar(value="Aguardando cálculo")
        self.var_pdf_state = tk.StringVar(value="PDF disponível" if REPORTLAB_OK else "PDF indisponível")
        self.var_excel_state = tk.StringVar(value="Excel disponível" if OPENPYXL_OK else "Excel indisponível")
        for var in (self.var_d, self.var_asx, self.var_asy, self.var_tipo_pilar, self.var_forma_pilar,
                    self.var_c1, self.var_c2, self.var_has_abertura, self.var_is_sapata):
            var.trace_add("write", self._on_geometry_change)
        for var in (self.var_d, self.var_asx, self.var_asy):
            var.trace_add("write", self._update_rho_label)

    def _build_ui(self):
        self.columnconfigure(0, weight=0)
        self.columnconfigure(1, weight=1)
        self.rowconfigure(0, weight=1)
        left = ttk.Frame(self, style="TFrame", padding=12)
        left.grid(row=0, column=0, sticky="nsew")
        left.rowconfigure(2, weight=1)
        right = ttk.Frame(self, style="TFrame", padding=(0, 12, 12, 12))
        right.grid(row=0, column=1, sticky="nsew")
        right.rowconfigure(2, weight=1)
        right.columnconfigure(0, weight=1)

        ttk.Label(left, text=APP_TITLE, style="Header.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 10))
        self._build_toolbar(left)
        self._build_left_notebook(left)
        self._build_right_panel(right)
        status = ttk.Label(self, textvariable=self.var_status, anchor="w", padding=(10, 6))
        status.grid(row=1, column=0, columnspan=2, sticky="ew")

    def _build_toolbar(self, parent):
        frm = ttk.Frame(parent, style="Card.TFrame", padding=10)
        frm.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        frm.columnconfigure(1, weight=1)
        ttk.Label(frm, text="Exemplo", style="Card.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Combobox(frm, textvariable=self.var_exemplo, values=list(EXEMPLOS.keys()), state="readonly").grid(row=0, column=1, sticky="ew", padx=(0, 8))
        ttk.Button(frm, text="Aplicar exemplo", command=self.carregar_exemplo).grid(row=0, column=2, padx=(0, 6))
        ttk.Button(frm, text="Calcular", command=self.calcular, style="Accent.TButton").grid(row=0, column=3, padx=(0, 6))
        ttk.Button(frm, text="Limpar", command=self.limpar).grid(row=0, column=4)

    def _build_left_notebook(self, parent):
        nb = ttk.Notebook(parent)
        nb.grid(row=2, column=0, sticky="nsew")
        tab_dados = self._make_scrollable_tab(nb, "Dados")
        tab_opcoes = self._make_scrollable_tab(nb, "Opções")
        tab_esquema = ttk.Frame(nb, padding=10)
        nb.add(tab_esquema, text="Geometria")
        self._build_tab_dados(tab_dados)
        self._build_tab_opcoes(tab_opcoes)
        self._build_tab_esquema(tab_esquema)

    def _make_scrollable_tab(self, notebook, title):
        outer = ttk.Frame(notebook)
        notebook.add(outer, text=title)
        outer.rowconfigure(0, weight=1)
        outer.columnconfigure(0, weight=1)
        canvas = tk.Canvas(outer, bg=BG, highlightthickness=0, bd=0)
        vsb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        inner = ttk.Frame(canvas, padding=10)
        window_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_inner_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_canvas_configure(event):
            canvas.itemconfigure(window_id, width=event.width)

        def _on_mousewheel(event):
            if canvas.winfo_height() < inner.winfo_reqheight():
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        inner.bind("<Configure>", _on_inner_configure)
        canvas.bind("<Configure>", _on_canvas_configure)
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        return inner

    def _build_tab_dados(self, parent):
        parent.columnconfigure(0, weight=1)
        frm_mat = ttk.LabelFrame(parent, text="Materiais", padding=10)
        frm_mat.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        self._add_labeled_entry(frm_mat, "fck (MPa)", self.var_fck, 0)
        self._add_labeled_entry(frm_mat, "fyk (MPa)", self.var_fyk, 1)
        self._add_labeled_entry(frm_mat, "fywk (MPa)", self.var_fywk, 2)

        frm_laje = ttk.LabelFrame(parent, text="Laje", padding=10)
        frm_laje.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        self._add_labeled_entry(frm_laje, "Altura útil d (m)", self.var_d, 0)
        self._add_labeled_entry(frm_laje, "As,lx (cm²/m)", self.var_asx, 1)
        self._add_labeled_entry(frm_laje, "As,ly (cm²/m)", self.var_asy, 2)
        self._add_labeled_entry(frm_laje, "σcp (MPa)", self.var_sigma_cp, 3)
        ttk.Label(frm_laje, textvariable=self.var_rho_calc, style="Muted.TLabel").grid(row=4, column=0, columnspan=2, sticky="w", pady=(6, 0))

        frm_pilar = ttk.LabelFrame(parent, text="Pilar", padding=10)
        frm_pilar.grid(row=2, column=0, sticky="ew", pady=(0, 8))
        self.frm_pilar = frm_pilar
        ttk.Label(frm_pilar, text="Tipo", style="Card.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=2)
        ttk.Combobox(frm_pilar, textvariable=self.var_tipo_pilar, values=["interior", "bordo", "canto"], state="readonly").grid(row=0, column=1, sticky="ew", pady=2)
        ttk.Label(frm_pilar, text="Forma", style="Card.TLabel").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=2)
        ttk.Combobox(frm_pilar, textvariable=self.var_forma_pilar, values=["retangular", "circular"], state="readonly").grid(row=1, column=1, sticky="ew", pady=2)
        self.lbl_c1 = ttk.Label(frm_pilar, text="c1 (m)", style="Card.TLabel")
        self.lbl_c1.grid(row=2, column=0, sticky="w", padx=(0, 8), pady=2)
        self.ent_c1 = ttk.Entry(frm_pilar, textvariable=self.var_c1)
        self.ent_c1.grid(row=2, column=1, sticky="ew", pady=2)
        self.lbl_c2 = ttk.Label(frm_pilar, text="c2 (m)", style="Card.TLabel")
        self.lbl_c2.grid(row=3, column=0, sticky="w", padx=(0, 8), pady=2)
        self.ent_c2 = ttk.Entry(frm_pilar, textvariable=self.var_c2)
        self.ent_c2.grid(row=3, column=1, sticky="ew", pady=2)

        frm_esf = ttk.LabelFrame(parent, text="Esforços de cálculo (ELU)", padding=10)
        frm_esf.grid(row=3, column=0, sticky="ew", pady=(0, 8))
        self._add_labeled_entry(frm_esf, "VEd (kN)", self.var_ved, 0)
        self._add_labeled_entry(frm_esf, "MEdx (kN·m)", self.var_medx, 1)
        self._add_labeled_entry(frm_esf, "MEdy (kN·m)", self.var_medy, 2)
        ttk.Label(frm_esf, text="Convenção em planta: e_x = MEdy/VEd e e_y = MEdx/VEd.", style="Muted.TLabel").grid(row=3, column=0, columnspan=2, sticky="w", pady=(6, 0))

    def _build_tab_opcoes(self, parent):
        parent.columnconfigure(0, weight=1)
        frm_extra = ttk.LabelFrame(parent, text="Condições adicionais", padding=10)
        frm_extra.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        ttk.Checkbutton(frm_extra, text="Elemento é sapata", variable=self.var_is_sapata, command=self._apply_visibility_rules).grid(row=0, column=0, columnspan=2, sticky="w", pady=2)
        self.lbl_sigma_gd = ttk.Label(frm_extra, text="σgd (kPa)", style="Card.TLabel")
        self.lbl_sigma_gd.grid(row=1, column=0, sticky="w", padx=(0, 8), pady=2)
        self.ent_sigma_gd = ttk.Entry(frm_extra, textvariable=self.var_sigma_gd)
        self.ent_sigma_gd.grid(row=1, column=1, sticky="ew", pady=2)
        ttk.Checkbutton(frm_extra, text="Existem aberturas próximas", variable=self.var_has_abertura, command=self._apply_visibility_rules).grid(row=2, column=0, columnspan=2, sticky="w", pady=(8, 2))
        self.lbl_u1_inef = ttk.Label(frm_extra, text="u1 ineficaz (m)", style="Card.TLabel")
        self.lbl_u1_inef.grid(row=3, column=0, sticky="w", padx=(0, 8), pady=2)
        self.ent_u1_inef = ttk.Entry(frm_extra, textvariable=self.var_u1_inef)
        self.ent_u1_inef.grid(row=3, column=1, sticky="ew", pady=2)

        frm_beta = ttk.LabelFrame(parent, text="Modo de avaliação de β", padding=10)
        frm_beta.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        ttk.Radiobutton(frm_beta, text="Simplificado", value="simplificado", variable=self.var_beta).grid(row=0, column=0, sticky="w")
        ttk.Radiobutton(frm_beta, text="EC2", value="ec2", variable=self.var_beta).grid(row=1, column=0, sticky="w")
        ttk.Radiobutton(frm_beta, text="fib Model Code 2010", value="fib", variable=self.var_beta).grid(row=2, column=0, sticky="w")

        frm_orient = ttk.LabelFrame(parent, text="Direção da excentricidade para EC2", padding=10)
        frm_orient.grid(row=2, column=0, sticky="ew", pady=(0, 8))
        self.chk_edge_interior = ttk.Checkbutton(frm_orient, text="No pilar de bordo, a excentricidade perpendicular ao bordo é dirigida para o interior", variable=self.var_edge_interior)
        self.chk_edge_interior.grid(row=0, column=0, sticky="w")
        self.chk_corner_interior = ttk.Checkbutton(frm_orient, text="No pilar de canto, a excentricidade resultante é dirigida para o interior", variable=self.var_corner_interior)
        self.chk_corner_interior.grid(row=1, column=0, sticky="w", pady=(6, 0))

        frm_notes = ttk.LabelFrame(parent, text="Notas", padding=10)
        frm_notes.grid(row=3, column=0, sticky="nsew")
        parent.rowconfigure(3, weight=1)
        txt = (
            "• Para pilares retangulares de bordo, o programa usa a expressão (6.44), o W1 da expressão (6.45) e k com c1/(2c2).\n"
            "• Para pilares de canto com excentricidade para o interior, o programa usa β = u1/u1*.\n"
            "• Se a excentricidade relevante for para o exterior, o programa aplica a expressão geral do EC2.\n"
            f"• {self.var_pdf_state.get()} | {self.var_excel_state.get()}."
        )
        ttk.Label(frm_notes, text=txt, justify="left", style="Muted.TLabel").grid(row=0, column=0, sticky="nw")

    def _build_tab_esquema(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)
        ttk.Label(parent, text="Desenho técnico da zona de punçoamento e editor gráfico das dimensões", style="Header.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
        box = ttk.Frame(parent, style="Card.TFrame", padding=10)
        box.grid(row=1, column=0, sticky="nsew")
        box.columnconfigure(0, weight=1)
        box.rowconfigure(0, weight=1)
        self.canvas_scheme = tk.Canvas(box, bg="white", height=420, highlightthickness=1, highlightbackground="#cbd5e1")
        self.canvas_scheme.grid(row=0, column=0, sticky="nsew")
        self.canvas_scheme.bind("<Button-1>", self._on_canvas_press)
        self.canvas_scheme.bind("<B1-Motion>", self._on_canvas_drag)
        controls = ttk.Frame(box, style="Card.TFrame")
        controls.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        controls.columnconfigure(1, weight=1)
        controls.columnconfigure(3, weight=1)
        controls.columnconfigure(5, weight=1)
        ttk.Label(controls, text="c1/D", style="Card.TLabel").grid(row=0, column=0, sticky="w")
        self.scale_c1 = tk.Scale(controls, from_=0.20, to=1.50, resolution=0.01, orient="horizontal", bg=CARD, highlightthickness=0, command=lambda v: self._set_var_from_scale(self.var_c1, v))
        self.scale_c1.grid(row=0, column=1, sticky="ew", padx=(6, 18))
        ttk.Label(controls, text="c2", style="Card.TLabel").grid(row=0, column=2, sticky="w")
        self.scale_c2 = tk.Scale(controls, from_=0.20, to=1.50, resolution=0.01, orient="horizontal", bg=CARD, highlightthickness=0, command=lambda v: self._set_var_from_scale(self.var_c2, v))
        self.scale_c2.grid(row=0, column=3, sticky="ew", padx=(6, 18))
        ttk.Label(controls, text="d", style="Card.TLabel").grid(row=0, column=4, sticky="w")
        self.scale_d = tk.Scale(controls, from_=0.12, to=0.60, resolution=0.005, orient="horizontal", bg=CARD, highlightthickness=0, command=lambda v: self._set_var_from_scale(self.var_d, v))
        self.scale_d.grid(row=0, column=5, sticky="ew", padx=(6, 0))
        ttk.Label(box, text="Arraste os puxadores azuis para alterar c1 e c2; o anel vermelho representa u1 a 2d.", style="Muted.TLabel").grid(row=2, column=0, sticky="w", pady=(8, 0))

    def _build_right_panel(self, parent):
        top = ttk.Frame(parent, style="Card.TFrame", padding=10)
        top.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        top.columnconfigure(1, weight=1)
        ttk.Label(top, text="Estado", style="Card.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 8))
        ttk.Label(top, textvariable=self.var_resultado, style="Card.TLabel").grid(row=0, column=1, sticky="w")
        self.lbl_badge = ttk.Label(top, text="—", style="Card.TLabel")
        self.lbl_badge.grid(row=0, column=2, sticky="e")
        cols = ("parâmetro", "valor")
        self.tree_summary = ttk.Treeview(top, columns=cols, show="headings", height=8)
        self.tree_summary.grid(row=1, column=0, columnspan=3, sticky="ew", pady=(8, 0))
        self.tree_summary.heading("parâmetro", text="Parâmetro")
        self.tree_summary.heading("valor", text="Valor")
        self.tree_summary.column("parâmetro", width=220, anchor="w")
        self.tree_summary.column("valor", width=160, anchor="center")

        actions = ttk.Frame(parent, style="TFrame")
        actions.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        for i in range(2):
            actions.rowconfigure(i, weight=1)
        for i in range(2):
            actions.columnconfigure(i, weight=1)
        ttk.Button(actions, text="Guardar TXT", command=self.guardar_relatorio_txt).grid(row=0, column=0, sticky="ew", padx=(0, 6), pady=(0, 6))
        ttk.Button(actions, text="Exportar PDF", command=self.guardar_relatorio_pdf).grid(row=0, column=1, sticky="ew", padx=(0, 0), pady=(0, 6))
        ttk.Button(actions, text="Exportar Excel", command=self.guardar_relatorio_excel).grid(row=1, column=0, sticky="ew", padx=(0, 6))
        ttk.Button(actions, text="Copiar relatório", command=self.copiar_relatorio).grid(row=1, column=1, sticky="ew")

        nb = ttk.Notebook(parent)
        nb.grid(row=2, column=0, sticky="nsew")
        tab_rel = ttk.Frame(nb, padding=6)
        tab_diag = ttk.Frame(nb, padding=6)
        nb.add(tab_rel, text="Relatório técnico")
        nb.add(tab_diag, text="Diagnóstico")
        tab_rel.rowconfigure(0, weight=1)
        tab_rel.columnconfigure(0, weight=1)
        self.txt_output = tk.Text(tab_rel, wrap="word", font=("Consolas", 10), bg="#fbfdff", fg=TEXT)
        self.txt_output.grid(row=0, column=0, sticky="nsew")
        sc1 = ttk.Scrollbar(tab_rel, orient="vertical", command=self.txt_output.yview)
        sc1.grid(row=0, column=1, sticky="ns")
        self.txt_output.configure(yscrollcommand=sc1.set)
        tab_diag.rowconfigure(0, weight=1)
        tab_diag.columnconfigure(0, weight=1)
        self.txt_diag = tk.Text(tab_diag, wrap="word", font=("Segoe UI", 10), state="disabled", bg="#fbfdff", fg=TEXT)
        self.txt_diag.grid(row=0, column=0, sticky="nsew")

    @staticmethod
    def _add_labeled_entry(parent, label, variable, row):
        ttk.Label(parent, text=label, style="Card.TLabel").grid(row=row, column=0, sticky="w", padx=(0, 8), pady=2)
        ttk.Entry(parent, textvariable=variable).grid(row=row, column=1, sticky="ew", pady=2)

    def _parse_float(self, value, field_name, min_value=None):
        try:
            number = float(str(value).replace(",", ".").strip())
        except Exception as exc:
            raise ValueError(f"Valor inválido em '{field_name}'.") from exc
        if min_value is not None and number < min_value:
            raise ValueError(f"O campo '{field_name}' deve ser ≥ {min_value}.")
        return number

    def _safe_float(self, value, default):
        try:
            return float(str(value).replace(",", ".").strip())
        except Exception:
            return default

    def _apply_visibility_rules(self):
        forma = self.var_forma_pilar.get().lower()
        tipo = self.var_tipo_pilar.get().lower()
        self.lbl_c1.configure(text="Diâmetro D (m)" if forma == "circular" else ("c1 (m) ‖ bordo" if tipo == "bordo" else "c1 (m)"))
        if forma == "circular":
            self.lbl_c2.grid_remove(); self.ent_c2.grid_remove(); self.scale_c2.configure(state="disabled")
        else:
            self.lbl_c2.configure(text="c2 (m) ⟂ bordo" if tipo == "bordo" else "c2 (m)")
            self.lbl_c2.grid(); self.ent_c2.grid(); self.scale_c2.configure(state="normal")
        if self.var_is_sapata.get():
            self.lbl_sigma_gd.grid(); self.ent_sigma_gd.grid()
        else:
            self.lbl_sigma_gd.grid_remove(); self.ent_sigma_gd.grid_remove()
        if self.var_has_abertura.get():
            self.lbl_u1_inef.grid(); self.ent_u1_inef.grid()
        else:
            self.lbl_u1_inef.grid_remove(); self.ent_u1_inef.grid_remove()
        if tipo == "bordo":
            self.chk_edge_interior.state(["!disabled"])
            self.chk_corner_interior.state(["disabled"])
        elif tipo == "canto":
            self.chk_edge_interior.state(["disabled"])
            self.chk_corner_interior.state(["!disabled"])
        else:
            self.chk_edge_interior.state(["disabled"])
            self.chk_corner_interior.state(["disabled"])
        self._sync_scales()
        self._draw_scheme()

    def _sync_scales(self):
        self.scale_c1.set(self._safe_float(self.var_c1.get(), 0.40))
        self.scale_c2.set(self._safe_float(self.var_c2.get(), 0.40))
        self.scale_d.set(self._safe_float(self.var_d.get(), 0.22))

    def _update_rho_label(self, *_):
        try:
            d = self._parse_float(self.var_d.get(), "d", 1e-9)
            asx = self._parse_float(self.var_asx.get(), "As,lx", 0.0)
            asy = self._parse_float(self.var_asy.get(), "As,ly", 0.0)
            rho = min(math.sqrt(((asx / 10000) / d) * ((asy / 10000) / d)), 0.02) if asx > 0 and asy > 0 else None
            self.var_rho_calc.set(f"ρl = {rho * 100:.3f} %" if rho is not None else "ρl = -")
        except Exception:
            self.var_rho_calc.set("ρl = -")

    def _on_geometry_change(self, *_):
        self.after_idle(self._apply_visibility_rules)

    def _collect_inputs(self):
        forma = self.var_forma_pilar.get().lower()
        data = {
            "betão_fck": self._parse_float(self.var_fck.get(), "fck", 12.0),
            "aço_fyk": self._parse_float(self.var_fyk.get(), "fyk", 1.0),
            "aço_fywk": self._parse_float(self.var_fywk.get(), "fywk", 1.0),
            "laje_d": self._parse_float(self.var_d.get(), "d", 1e-6),
            "laje_As_lx_cm2pm": self._parse_float(self.var_asx.get(), "As,lx", 0.0),
            "laje_As_ly_cm2pm": self._parse_float(self.var_asy.get(), "As,ly", 0.0),
            "sigma_cp": self._parse_float(self.var_sigma_cp.get(), "σcp", 0.0),
            "pilar_tipo": self.var_tipo_pilar.get().lower(),
            "pilar_forma": forma,
            "V_Ed": self._parse_float(self.var_ved.get(), "VEd", 0.0) * 1000.0,
            "pilar_c1": self._parse_float(self.var_c1.get(), "c1 / D", 1e-6),
            "pilar_c2": None,
            "M_Edx": self._parse_float(self.var_medx.get(), "MEdx") * 1000.0,
            "M_Edy": self._parse_float(self.var_medy.get(), "MEdy") * 1000.0,
            "is_sapata": bool(self.var_is_sapata.get()),
            "sigma_gd_kpa": self._parse_float(self.var_sigma_gd.get(), "σgd", 0.0) if self.var_is_sapata.get() else 0.0,
            "u1_ineffective": self._parse_float(self.var_u1_inef.get(), "u1 ineficaz", 0.0) if self.var_has_abertura.get() else 0.0,
            "beta_mode": self.var_beta.get().lower(),
            "edge_perp_interior": bool(self.var_edge_interior.get()),
            "corner_interior": bool(self.var_corner_interior.get()),
        }
        if forma == "retangular":
            data["pilar_c2"] = self._parse_float(self.var_c2.get(), "c2", 1e-6)
        return data

    def calcular(self):
        try:
            inputs = self._collect_inputs()
            verif = PuncoamentoEC2(**inputs)
            report = verif.verificar_puncoamento()
            self.last_verif = verif
            self.last_report = report
            self.txt_output.delete("1.0", tk.END)
            self.txt_output.insert(tk.END, report)
            self._fill_summary(verif)
            self._fill_diagnostic(verif)
            self._draw_scheme(verif)
            self.var_status.set("Cálculo concluído com sucesso.")
        except Exception as exc:
            self.var_status.set("Erro no cálculo.")
            messagebox.showerror("Erro", str(exc))

    def _fill_summary(self, verif):
        for item in self.tree_summary.get_children():
            self.tree_summary.delete(item)
        rows = [
            ("β", f"{verif.beta:.3f}"), ("u0", f"{verif.u0:.3f} m"), ("u1", f"{verif.u1:.3f} m"),
            ("u1,ef", f"{verif.u1_eff:.3f} m"), ("vEd(u0)", f"{verif.v_Ed_u0:.3f} MPa"),
            ("vRd,max", f"{verif.v_Rd_max:.3f} MPa"), ("vEd(u1)", f"{verif.v_Ed_u1:.3f} MPa"),
            ("vRd,c", f"{verif.v_Rd_c:.3f} MPa"),
        ]
        for r in rows:
            self.tree_summary.insert("", tk.END, values=r)
        ok = (verif.v_Ed_u0 <= verif.v_Rd_max) and (verif.v_Ed_u1 <= verif.v_Rd_c)
        self.var_resultado.set("Não é necessária armadura de punçoamento" if ok else ("É necessária armadura de punçoamento" if verif.v_Ed_u0 <= verif.v_Rd_max else "Falha no perímetro u0"))
        self.lbl_badge.configure(text="OK" if ok else ("ATENÇÃO" if verif.v_Ed_u0 <= verif.v_Rd_max else "FALHA"), foreground=(OK if ok else (WARN if verif.v_Ed_u0 <= verif.v_Rd_max else FAIL)))

    def _fill_diagnostic(self, verif):
        ratio_u0 = verif.v_Ed_u0 / verif.v_Rd_max if verif.v_Rd_max else float("inf")
        ratio_u1 = verif.v_Ed_u1 / verif.v_Rd_c if verif.v_Rd_c else float("inf")
        lines = [
            "Leitura rápida do resultado\n",
            f"• Utilização em u0 = {ratio_u0:.3f}",
            f"• Utilização em u1 = {ratio_u1:.3f}",
            f"• Modo de β = {self.var_beta.get()}",
            "",
        ]
        if ratio_u0 > 1.0:
            lines += ["A verificação em u0 falha.", "Reveja d, fck ou a geometria da zona de apoio."]
        elif ratio_u1 > 1.0:
            lines += ["A verificação sem armadura de punçoamento não satisfaz.", "É necessário prever armadura específica de punçoamento."]
        else:
            lines += ["A verificação sem armadura de punçoamento satisfaz."]
        if self.var_tipo_pilar.get() == "bordo":
            lines.append("Para bordo, o EC2 usa u1*, a expressão (6.44) e W1 pela expressão (6.45) quando a excentricidade perpendicular é interior.")
        if self.var_tipo_pilar.get() == "canto":
            lines.append("Para canto, com excentricidade para o interior, o EC2 admite β = u1/u1*.")
        self.txt_diag.configure(state="normal")
        self.txt_diag.delete("1.0", tk.END)
        self.txt_diag.insert(tk.END, "\n".join(lines))
        self.txt_diag.configure(state="disabled")

    def _set_var_from_scale(self, var, value):
        var.set(f"{float(value):.3f}")

    def _draw_scheme(self, verif=None):
        if not hasattr(self, "canvas_scheme"):
            return
        cv = self.canvas_scheme
        cv.delete("all")
        w = max(cv.winfo_width(), 700)
        h = max(cv.winfo_height(), 420)
        pad = 35
        slab_left, slab_top, slab_right, slab_bottom = pad, pad, w - pad, h - pad
        cv.create_rectangle(slab_left, slab_top, slab_right, slab_bottom, outline="#cbd5e1", width=1, dash=(6, 4))
        cv.create_text(slab_left + 10, slab_top + 12, anchor="w", text="Contorno da laje / zona representativa", fill=MUTED, font=("Segoe UI", 9))
        forma = self.var_forma_pilar.get().lower()
        tipo = self.var_tipo_pilar.get().lower()
        c1 = self._safe_float(self.var_c1.get(), 0.40)
        c2 = self._safe_float(self.var_c2.get(), 0.40) if forma == "retangular" else self._safe_float(self.var_c1.get(), 0.40)
        d = self._safe_float(self.var_d.get(), 0.22)
        scale = min((w - 160) / max(c1 + 8 * d, c2 + 8 * d, 1.0), (h - 110) / max(c1 + 8 * d, c2 + 8 * d, 1.0))
        scale = max(scale, 120)
        pw, ph = c1 * scale, c2 * scale
        cx, cy = (slab_left + slab_right) / 2, (slab_top + slab_bottom) / 2
        if tipo == "bordo":
            cy = slab_top + ph / 2
        elif tipo == "canto":
            cx, cy = slab_left + pw / 2, slab_top + ph / 2
        px1, py1, px2, py2 = cx - pw / 2, cy - ph / 2, cx + pw / 2, cy + ph / 2
        # pilar
        if forma == "retangular":
            cv.create_rectangle(px1, py1, px2, py2, fill="#dbeafe", outline=ACCENT, width=2)
        else:
            cv.create_oval(px1, py1, px2, py2, fill="#dbeafe", outline=ACCENT, width=2)
        cv.create_text(cx, cy, text="Pilar", font=("Segoe UI", 10, "bold"), fill=TEXT)
        # u0 técnico
        self._draw_u0(cv, tipo, forma, px1, py1, px2, py2, d * scale)
        # u1 técnico
        self._draw_u1(cv, tipo, forma, px1, py1, px2, py2, 2 * d * scale)
        # handles
        if forma == "retangular":
            self.handle_c1 = (px2, cy)
            self.handle_c2 = (cx, py2)
            cv.create_oval(px2 - 6, cy - 6, px2 + 6, cy + 6, fill=ACCENT, outline=ACCENT, tags=("handle_c1",))
            cv.create_oval(cx - 6, py2 - 6, cx + 6, py2 + 6, fill=ACCENT, outline=ACCENT, tags=("handle_c2",))
        else:
            self.handle_c1 = (px2, cy)
            self.handle_c2 = None
            cv.create_oval(px2 - 6, cy - 6, px2 + 6, cy + 6, fill=ACCENT, outline=ACCENT, tags=("handle_c1",))
        # cotas
        self._dim_line(cv, px1, py2 + 24, px2, py2 + 24, f"c1 = {c1:.3f} m")
        if forma == "retangular":
            self._dim_line(cv, px2 + 24, py1, px2 + 24, py2, f"c2 = {c2:.3f} m", vertical=True)
        self._dim_line(cv, px2 + 62, py2 - 2 * d * scale, px2 + 62, py2, f"2d = {2*d:.3f} m", vertical=True, color="#dc2626")
        if self.var_has_abertura.get():
            ax, ay = px2 + 90, py1 + 20
            cv.create_rectangle(ax, ay, ax + 64, ay + 28, outline="#f59e0b", fill="#fef3c7")
            cv.create_text(ax + 32, ay + 14, text="Abertura", fill=WARN, font=("Segoe UI", 8, "bold"))
            cv.create_line(px2 + 2 * d * scale, cy, ax, ay + 14, arrow="last", fill="#f59e0b")
        if verif is not None:
            cv.create_text(slab_left + 10, slab_bottom - 10, anchor="sw", text=f"β = {verif.beta:.3f} | u0 = {verif.u0:.3f} m | u1 = {verif.u1:.3f} m", fill=TEXT, font=("Segoe UI", 10, "bold"))
        legend = "u0 junto ao pilar" if tipo == "interior" else "u0 limitado por bordo livre"
        cv.create_text(slab_left + 10, slab_top + 32, anchor="w", text=f"Azul: pilar | Roxo: u0 | Vermelho: u1 a 2d | {legend}", fill=MUTED, font=("Segoe UI", 9))

    def _draw_u0(self, cv, tipo, forma, x1, y1, x2, y2, doff):
        col = "#7c3aed"
        if tipo == "interior":
            if forma == "retangular":
                cv.create_rectangle(x1, y1, x2, y2, outline=col, width=2)
            else:
                cv.create_oval(x1, y1, x2, y2, outline=col, width=2)
            cv.create_text(x2 + 10, y1 - 8, anchor="w", text="u0", fill=col, font=("Segoe UI", 9, "bold"))
        elif tipo == "bordo":
            cv.create_line(x1, y2 + doff * 1.5, x2, y2 + doff * 1.5, fill=col, width=2)
            cv.create_line(x2, y1, x2, y2 + doff * 1.5, fill=col, width=2)
            cv.create_arc(x1 - 2 * doff, y2 - doff * 0.5, x1 + doff, y2 + doff * 2.0, start=90, extent=90, style="arc", outline=col, width=2)
            cv.create_text(x2 + 10, y2 + doff * 1.5, anchor="w", text="u0", fill=col, font=("Segoe UI", 9, "bold"))
        else:
            cv.create_arc(x2 - doff * 1.5, y2 - doff * 1.5, x2 + doff * 1.5, y2 + doff * 1.5, start=180, extent=90, style="arc", outline=col, width=2)
            cv.create_text(x2 + 10, y2 + 10, anchor="w", text="u0", fill=col, font=("Segoe UI", 9, "bold"))

    def _draw_u1(self, cv, tipo, forma, x1, y1, x2, y2, off):
        col = "#dc2626"
        if forma == "circular":
            if tipo == "interior":
                cv.create_oval(x1 - off, y1 - off, x2 + off, y2 + off, outline=col, width=2)
            elif tipo == "bordo":
                cv.create_arc(x1 - off, y1 - off, x2 + off, y2 + off, start=180, extent=180, style="arc", outline=col, width=2)
                cv.create_line(x1 - off, y2, x2 + off, y2, fill=col, width=2)
            else:
                cv.create_arc(x1 - off, y1 - off, x2 + off, y2 + off, start=180, extent=90, style="arc", outline=col, width=2)
                cv.create_line(x1, y2 + off, x2 + off * 0.5, y2 + off, fill=col, width=2)
                cv.create_line(x2 + off, y1, x2 + off, y2 + off * 0.5, fill=col, width=2)
        else:
            if tipo == "interior":
                cv.create_rectangle(x1 - off, y1 - off, x2 + off, y2 + off, outline=col, width=2)
            elif tipo == "bordo":
                cv.create_line(x1 - off, y2 + off, x2 + off, y2 + off, fill=col, width=2)
                cv.create_line(x2 + off, y1, x2 + off, y2 + off, fill=col, width=2)
                cv.create_line(x1 - off, y1, x1 - off, y2 + off, fill=col, width=2)
                cv.create_arc(x1 - 2*off, y2, x1, y2 + 2*off, start=90, extent=90, style="arc", outline=col, width=2)
                cv.create_arc(x2, y2, x2 + 2*off, y2 + 2*off, start=0, extent=90, style="arc", outline=col, width=2)
            else:
                cv.create_line(x1, y2 + off, x2 + off, y2 + off, fill=col, width=2)
                cv.create_line(x2 + off, y1, x2 + off, y2 + off, fill=col, width=2)
                cv.create_arc(x2, y2, x2 + 2*off, y2 + 2*off, start=180, extent=90, style="arc", outline=col, width=2)
        cv.create_text(x2 + off + 10, y1 - off, anchor="w", text="u1", fill=col, font=("Segoe UI", 9, "bold"))

    def _dim_line(self, cv, x1, y1, x2, y2, text, vertical=False, color="#334155"):
        cv.create_line(x1, y1, x2, y2, fill=color)
        if vertical:
            cv.create_line(x1 - 6, y1, x1 + 6, y1, fill=color)
            cv.create_line(x2 - 6, y2, x2 + 6, y2, fill=color)
            cv.create_text(x1 + 8, (y1 + y2) / 2, anchor="w", text=text, fill=color, font=("Segoe UI", 9))
        else:
            cv.create_line(x1, y1 - 6, x1, y1 + 6, fill=color)
            cv.create_line(x2, y2 - 6, x2, y2 + 6, fill=color)
            cv.create_text((x1 + x2) / 2, y1 - 12, text=text, fill=color, font=("Segoe UI", 9))

    def _on_canvas_press(self, event):
        item = self.canvas_scheme.find_closest(event.x, event.y)
        tags = self.canvas_scheme.gettags(item)
        self.drag_mode = "c1" if "handle_c1" in tags else ("c2" if "handle_c2" in tags else None)

    def _on_canvas_drag(self, event):
        if self.drag_mode is None or self.var_forma_pilar.get().lower() not in ("retangular", "circular"):
            return
        w = max(self.canvas_scheme.winfo_width(), 700)
        h = max(self.canvas_scheme.winfo_height(), 420)
        c1 = self._safe_float(self.var_c1.get(), 0.40)
        c2 = self._safe_float(self.var_c2.get(), 0.40)
        d = self._safe_float(self.var_d.get(), 0.22)
        scale = min((w - 160) / max(c1 + 8 * d, c2 + 8 * d, 1.0), (h - 110) / max(c1 + 8 * d, c2 + 8 * d, 1.0))
        scale = max(scale, 120)
        cx, cy = (35 + w - 35) / 2, (35 + h - 35) / 2
        if self.var_tipo_pilar.get() == "bordo":
            cy = 35 + c2 * scale / 2
        elif self.var_tipo_pilar.get() == "canto":
            cx, cy = 35 + c1 * scale / 2, 35 + c2 * scale / 2
        if self.drag_mode == "c1":
            new_c1 = max(0.20, min(1.50, 2 * abs(event.x - cx) / scale))
            self.var_c1.set(f"{new_c1:.3f}")
            if self.var_forma_pilar.get().lower() == "circular":
                self.var_c2.set(f"{new_c1:.3f}")
        elif self.drag_mode == "c2" and self.var_forma_pilar.get().lower() == "retangular":
            new_c2 = max(0.20, min(1.50, 2 * abs(event.y - cy) / scale))
            self.var_c2.set(f"{new_c2:.3f}")
        self._draw_scheme()

    def guardar_relatorio_txt(self):
        content = self.txt_output.get("1.0", tk.END).strip()
        if not content:
            messagebox.showinfo("Guardar relatório", "Não existe relatório para guardar.")
            return
        filepath = filedialog.asksaveasfilename(title="Guardar relatório TXT", defaultextension=".txt", filetypes=[("Texto", "*.txt")])
        if not filepath:
            return
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(content)
        self.var_status.set(f"Relatório TXT guardado em: {filepath}")

    def guardar_relatorio_pdf(self):
        content = self.txt_output.get("1.0", tk.END).strip()
        if not content:
            messagebox.showinfo("Exportar PDF", "Não existe relatório para exportar.")
            return
        if not REPORTLAB_OK:
            messagebox.showerror("Exportar PDF", "A biblioteca ReportLab não está disponível neste ambiente.")
            return
        filepath = filedialog.asksaveasfilename(title="Exportar relatório PDF", defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
        if not filepath:
            return
        self._create_pdf(filepath, content)
        self.var_status.set(f"Relatório PDF guardado em: {filepath}")

    def _create_pdf(self, filepath, content):
        emitted_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        c = pdf_canvas.Canvas(filepath, pagesize=A4)
        width, height = A4
        x0 = 2.0 * cm
        top_margin = 2.2 * cm
        bottom_margin = 1.8 * cm
        usable_w = width - 4.0 * cm
        body_font = "Courier"
        body_bold = "Courier-Bold"
        body_size = 10
        subtitle_size = 12
        footer_size = 8
        body_leading = body_size * 1.5
        section_gap = body_size * 2.0
        subtitle_gap = body_size * 2.0
        repo_url = "https://github.com/lutondatomalela/PunchingShearEC2"

        sections = self._build_professional_report_sections(emitted_at)
        first_page = True

        def draw_footer():
            c.setFont(body_font, footer_size)
            footer = f"PunchingShearEC2 | {emitted_at}"
            footer_w = stringWidth(footer, body_font, footer_size)
            footer_x = (width - footer_w) / 2
            footer_y = 1.0 * cm
            c.drawString(footer_x, footer_y, footer)
            c.linkURL(repo_url, (footer_x, footer_y-2, footer_x + footer_w, footer_y + footer_size), relative=0)

        def new_page():
            nonlocal first_page
            if not first_page:
                c.showPage()
            y = height - top_margin
            if first_page:
                c.setTitle("PunchingShearEC2 - Relatório de Verificação de Punçoamento")
                c.setFont(body_bold, 14)
                c.drawString(x0, y, "Relatório de verificação de punçoamento")
                y -= 0.9 * cm
                c.setFont(body_font, body_size)
                prog_text = "Programa: PunchingShearEC2"
                c.drawString(x0, y, prog_text)
                c.linkURL("https://github.com/lutondatomalela/PunchingShearEC2", (x0, y-2, x0 + stringWidth(prog_text, body_font, body_size), y + body_size), relative=0)
                y -= 0.55 * cm
                c.drawString(x0, y, "Norma de referência principal: NP EN 1992-1-1")
                y -= 0.8 * cm
                first_page = False
            return y

        y = new_page()
        c.setFont(body_font, body_size)
        for section in sections:
            title = section.get("title", "")
            lines = section.get("lines", [])
            if title:
                needed = subtitle_size + subtitle_gap + body_leading * max(1, len(lines)) + section_gap
                if y < bottom_margin + needed:
                    draw_footer()
                    y = new_page()
                c.setFont(body_bold, subtitle_size)
                c.drawString(x0, y, title)
                y -= subtitle_gap
            c.setFont(body_font, body_size)
            for line in lines:
                wrapped_lines = self._wrap_text(line if line else " ", body_font, body_size, usable_w)
                for wrapped in wrapped_lines:
                    if y < bottom_margin + body_leading:
                        draw_footer()
                        y = new_page()
                        c.setFont(body_font, body_size)
                    c.drawString(x0, y, wrapped)
                    y -= body_leading
            y -= section_gap

        draw_footer()
        c.save()

    def _build_professional_report_sections(self, emitted_at):
        v = self.last_verif
        if v is None:
            return [{"title": "", "lines": [self.txt_output.get("1.0", tk.END).strip()]}]

        def fmt(num, nd=3):
            try:
                return f"{float(num):.{nd}f}"
            except Exception:
                return "-"

        def yesno(flag):
            return "Sim" if flag else "Não"

        tipo = self.var_tipo_pilar.get().capitalize()
        forma = self.var_forma_pilar.get().capitalize()
        c2_label = "Diâmetro do pilar D [m]" if self.var_forma_pilar.get().lower() == "circular" else "Dimensão do pilar c2 [m]"
        ratio_u0 = v.v_Ed_u0 / v.v_Rd_max if getattr(v, 'v_Rd_max', 0) else float('inf')
        ratio_u1 = v.v_Ed_u1 / v.v_Rd_c if getattr(v, 'v_Rd_c', 0) else float('inf')

        sections = [
            {"title": "1. Info", "lines": [
                "Verificação de punçoamento de ligação laje-pilar segundo a NP EN 1992-1-1, com avaliação da resistência na face do pilar e da resistência sem armadura de punçoamento no perímetro de controlo básico.",
            ]},
            {"title": "2. Dados de entrada", "lines": [
                f"Resistência característica do betão fck [MPa]: {fmt(v.fck)}",
                f"Tensão característica de cedência das armaduras longitudinais fyk [MPa]: {fmt(v.fyk)}",
                f"Tensão característica de cedência das armaduras de punçoamento fywk [MPa]: {fmt(v.fywk)}",
                f"Altura útil da laje d [m]: {fmt(v.d)}",
                f"Armadura longitudinal As,lx [cm²/m]: {self.var_asx.get()}",
                f"Armadura longitudinal As,ly [cm²/m]: {self.var_asy.get()}",
                f"Taxa média de armadura ρl [%]: {fmt(v.rho_l * 100)}",
                f"Tensão média de compressão σcp [MPa]: {fmt(v.sigma_cp)}",
                f"Tipo de pilar: {tipo}",
                f"Forma do pilar: {forma}",
                f"Dimensão do pilar c1 [m]: {fmt(v.c1)}",
                f"{c2_label}: {fmt(v.c2 if getattr(v,'c2',None) is not None else getattr(v,'D',None))}",
                f"Esforço transverso de cálculo VEd [kN]: {fmt(v.V_Ed / 1000)}",
                f"Momento fletor de cálculo MEdx [kN·m]: {fmt(v.M_Edx / 1000)}",
                f"Momento fletor de cálculo MEdy [kN·m]: {fmt(v.M_Edy / 1000)}",
                f"Elemento de fundação tipo sapata: {yesno(v.is_sapata)}",
                f"Existem aberturas próximas consideradas na redução de perímetro: {yesno(v.u1_ineffective > 0)}",
                f"Perímetro ineficaz devido a aberturas u1,inef [m]: {fmt(v.u1_ineffective)}",
                f"Modo de avaliação do coeficiente β: {self.var_beta.get().upper()}",
            ]},
            {"title": "3. Parâmetros geométricos e mecânicos", "lines": [
                f"Perímetro na face do pilar u0 [m]: {fmt(v.u0)}",
                f"Perímetro de controlo básico u1 [m]: {fmt(v.u1)}",
                f"Perímetro de controlo efetivo u1,ef [m]: {fmt(v.u1_eff)}",
                f"Esforço transverso reduzido VEd,red [kN]: {fmt(v.V_Ed_red / 1000)}",
                f"Coeficiente β [-]: {fmt(v.beta)}",
                f"Coeficiente k [-]: {fmt(getattr(v, 'k_val', None))}",
                f"Tensão de cálculo v_Ed(u0) [MPa]: {fmt(v.v_Ed_u0)}",
                f"Tensão resistente máxima v_Rd,max [MPa]: {fmt(v.v_Rd_max)}",
                f"Tensão de cálculo v_Ed(u1) [MPa]: {fmt(v.v_Ed_u1)}",
                f"Tensão resistente do betão v_Rd,c [MPa]: {fmt(v.v_Rd_c)}",
                f"Índice de utilização em u0 [-]: {fmt(ratio_u0)}",
                f"Índice de utilização em u1 [-]: {fmt(ratio_u1)}",
            ]},
            {"title": "4. Referências normativas adotadas", "lines": [
                "Determinação do perímetro de controlo básico conforme 6.4.2 da NP EN 1992-1-1.",
                "Avaliação do coeficiente β conforme 6.4.3, incluindo expressões (6.39), (6.41), (6.44), (6.45) e (6.46), conforme aplicável ao tipo de pilar e à direção da excentricidade.",
                "Resistência ao punçoamento sem armadura conforme 6.4.4, com v_Rd,c pela expressão (6.47).",
                "Resistência máxima junto ao pilar conforme 6.4.5 e respetiva nota nacional adotada no programa para v_Rd,max.",
                "Dimensionamento da armadura de punçoamento, quando necessária, conforme 6.4.5 e expressão (6.52).",
            ]},
            {"title": "5. Verificações realizadas", "lines": [
                f"Verificação na face do pilar: v_Ed(u0) = {fmt(v.v_Ed_u0)} MPa {'≤' if v.v_Ed_u0 <= v.v_Rd_max else '>'} v_Rd,max = {fmt(v.v_Rd_max)} MPa.",
                f"Verificação no perímetro de controlo básico: v_Ed(u1) = {fmt(v.v_Ed_u1)} MPa {'≤' if v.v_Ed_u1 <= v.v_Rd_c else '>'} v_Rd,c = {fmt(v.v_Rd_c)} MPa.",
            ]},
            {"title": "6. Conclusão", "lines": [
                self.var_resultado.get() + ".",
            ]},
        ]

        det_lines = []
        if getattr(v, 'armadura_necessaria', False):
            det_lines.extend([
                f"Tensão resistente máxima com armadura v_Rd,cs,max [MPa]: {fmt(getattr(v, 'v_Rd_cs_max', None))}",
                f"Tensão efetiva de cálculo do aço de punçoamento f_ywd,ef [MPa]: {fmt(getattr(v, 'f_ywd_ef', None))}",
                f"Armadura calculada Asw/sr [cm²/m]: {fmt(getattr(v, 'Asw_sr_calc', 0) * 1e4)}",
                f"Armadura mínima Asw/sr [cm²/m]: {fmt(getattr(v, 'Asw_sr_min', 0) * 1e4)}",
                f"Armadura adotada Asw/sr [cm²/m]: {fmt(getattr(v, 'Asw_sr_req', 0) * 1e4)}",
                f"Perímetro exterior efetivo u_out,ef [m]: {fmt(getattr(v, 'u_out_ef', None))}",
                f"Extensão radial a armar a partir da face do pilar [m]: {fmt(getattr(v, 'dist_zona_armar', None))}",
                f"Posição máxima do primeiro perímetro s0,max [m]: {fmt(getattr(v, 's0_max', None))}",
                f"Espaçamento radial máximo entre perímetros sr,max [m]: {fmt(getattr(v, 'sr_max', None))}",
                f"Número estimado de perímetros de armadura: {getattr(v, 'n_perimetros', '-')}",
                f"Área de armadura por perímetro Asw [cm²]: {fmt(getattr(v, 'Asw_por_perimetro', 0) * 1e4)}",
                "Pormenorização resultante: dispor o primeiro perímetro a uma distância não superior a 0,5d da face do pilar e os perímetros seguintes com espaçamento radial não superior a 0,75d, prolongando a armadura até ao limite definido por u_out,ef.",
            ])
        else:
            det_lines.append("Pormenorização resultante: não é necessária armadura específica de punçoamento para a situação verificada.")
        sections.append({"title": "7. Recomendação de pormenorização", "lines": det_lines})

        note_lines = []
        if self.var_tipo_pilar.get() == 'bordo':
            note_lines.append("Nota: para pilar de bordo, o programa adota a formulação do EC2 com u1*, β pela expressão (6.44), quando aplicável, e W1 pela expressão (6.45).")
        elif self.var_tipo_pilar.get() == 'canto':
            note_lines.append("Nota: para pilar de canto com excentricidade dirigida para o interior da laje, o programa adota β = u1/u1* conforme a expressão (6.46).")
        if note_lines:
            sections.append({"title": "", "lines": note_lines})

        return sections

    def guardar_relatorio_excel(self):
        if not OPENPYXL_OK:
            messagebox.showerror("Exportar Excel", "A biblioteca openpyxl não está disponível neste ambiente.")
            return
        if self.last_verif is None:
            messagebox.showinfo("Exportar Excel", "Não existe cálculo concluído para exportar.")
            return
        filepath = filedialog.asksaveasfilename(title="Exportar relatório Excel", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not filepath:
            return
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Resumo"
        ws2 = wb.create_sheet("Relatorio")
        ws3 = wb.create_sheet("Entradas")
        bold = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="1D4ED8")
        thin = Side(style="thin", color="CBD5E1")
        border = Border(bottom=thin)
        # resumo
        ws1["A1"] = "Verificação de punçoamento"
        ws1["A1"].font = Font(bold=True, size=14)
        ws1["A3"] = "Parâmetro"; ws1["B3"] = "Valor"
        for c in ("A3", "B3"):
            ws1[c].font = bold; ws1[c].fill = fill; ws1[c].alignment = Alignment(horizontal="center")
        rows = [
            ("Resultado", self.var_resultado.get()), ("β", self.last_verif.beta), ("u0 (m)", self.last_verif.u0),
            ("u1 (m)", self.last_verif.u1), ("u1,ef (m)", self.last_verif.u1_eff), ("vEd(u0) (MPa)", self.last_verif.v_Ed_u0),
            ("vRd,max (MPa)", self.last_verif.v_Rd_max), ("vEd(u1) (MPa)", self.last_verif.v_Ed_u1), ("vRd,c (MPa)", self.last_verif.v_Rd_c),
        ]
        for i, (k, v) in enumerate(rows, start=4):
            ws1[f"A{i}"] = k; ws1[f"B{i}"] = v
            ws1[f"A{i}"].border = border; ws1[f"B{i}"].border = border
        ws1.column_dimensions["A"].width = 24
        ws1.column_dimensions["B"].width = 24
        # relatorio
        ws2["A1"] = "Relatório técnico"
        ws2["A1"].font = Font(bold=True, size=14)
        for i, line in enumerate(self.last_report.splitlines(), start=3):
            ws2[f"A{i}"] = line
        ws2.column_dimensions["A"].width = 120
        # entradas
        ws3["A1"] = "Entradas do cálculo"; ws3["A1"].font = Font(bold=True, size=14)
        ws3["A3"] = "Campo"; ws3["B3"] = "Valor"
        for c in ("A3", "B3"):
            ws3[c].font = bold; ws3[c].fill = fill
        entries = [
            ("fck (MPa)", self.var_fck.get()), ("fyk (MPa)", self.var_fyk.get()), ("fywk (MPa)", self.var_fywk.get()),
            ("d (m)", self.var_d.get()), ("As,lx (cm²/m)", self.var_asx.get()), ("As,ly (cm²/m)", self.var_asy.get()),
            ("σcp (MPa)", self.var_sigma_cp.get()), ("Tipo", self.var_tipo_pilar.get()), ("Forma", self.var_forma_pilar.get()),
            ("c1/D (m)", self.var_c1.get()), ("c2 (m)", self.var_c2.get()), ("VEd (kN)", self.var_ved.get()),
            ("MEdx (kN·m)", self.var_medx.get()), ("MEdy (kN·m)", self.var_medy.get()), ("Modo β", self.var_beta.get()),
        ]
        for i, (k, v) in enumerate(entries, start=4):
            ws3[f"A{i}"] = k; ws3[f"B{i}"] = v
        ws3.column_dimensions["A"].width = 28
        ws3.column_dimensions["B"].width = 20
        for ws in (ws1, ws2, ws3):
            ws.freeze_panes = "A4"
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].bestFit = True
        wb.save(filepath)
        self.var_status.set(f"Relatório Excel guardado em: {filepath}")

    @staticmethod
    def _wrap_text(text, font_name, font_size, max_width):
        words = text.split()
        if not words:
            return [""]
        lines, current = [], words[0]
        for word in words[1:]:
            test = current + " " + word
            if stringWidth(test, font_name, font_size) <= max_width:
                current = test
            else:
                lines.append(current); current = word
        lines.append(current)
        return lines

    def copiar_relatorio(self):
        content = self.txt_output.get("1.0", tk.END).strip()
        if not content:
            messagebox.showinfo("Copiar relatório", "Não existe relatório para copiar.")
            return
        self.clipboard_clear(); self.clipboard_append(content)
        self.var_status.set("Relatório copiado para a área de transferência.")

    def limpar(self):
        self.__init_reset__()

    def __init_reset__(self):
        vals = {
            self.var_fck: "30", self.var_fyk: "500", self.var_fywk: "500", self.var_d: "0.22", self.var_asx: "10.0", self.var_asy: "10.0",
            self.var_sigma_cp: "0", self.var_tipo_pilar: "interior", self.var_forma_pilar: "retangular", self.var_c1: "0.40", self.var_c2: "0.40",
            self.var_ved: "600", self.var_medx: "0", self.var_medy: "0", self.var_sigma_gd: "150", self.var_u1_inef: "0", self.var_beta: "simplificado"
        }
        for var, val in vals.items(): var.set(val)
        self.var_is_sapata.set(False); self.var_has_abertura.set(False); self.var_edge_interior.set(True); self.var_corner_interior.set(True)
        self.last_report = ""; self.last_verif = None
        self.txt_output.delete("1.0", tk.END)
        self.txt_diag.configure(state="normal"); self.txt_diag.delete("1.0", tk.END); self.txt_diag.configure(state="disabled")
        for item in self.tree_summary.get_children(): self.tree_summary.delete(item)
        self.var_resultado.set("Aguardando cálculo")
        self.var_status.set("Campos repostos.")
        self._apply_visibility_rules(); self._update_rho_label(); self._draw_scheme()

    def carregar_exemplo(self):
        ex = EXEMPLOS.get(self.var_exemplo.get())
        if not ex:
            return
        self.var_fck.set(ex["fck"]); self.var_fyk.set(ex["fyk"]); self.var_fywk.set(ex["fywk"])
        self.var_d.set(ex["d"]); self.var_asx.set(ex["asx"]); self.var_asy.set(ex["asy"]); self.var_sigma_cp.set(ex["sigma_cp"])
        self.var_tipo_pilar.set(ex["tipo"]); self.var_forma_pilar.set(ex["forma"]); self.var_c1.set(ex["c1"]); self.var_c2.set(ex["c2"])
        self.var_ved.set(ex["ved"]); self.var_medx.set(ex["medx"]); self.var_medy.set(ex["medy"])
        self.var_is_sapata.set(ex["is_sapata"]); self.var_sigma_gd.set(ex["sigma_gd"]); self.var_has_abertura.set(ex["has_abertura"])
        self.var_u1_inef.set(ex["u1_inef"]); self.var_beta.set(ex["beta"]); self.var_edge_interior.set(ex["edge_interior"]); self.var_corner_interior.set(ex["corner_interior"])
        self.var_status.set(f"Exemplo '{self.var_exemplo.get()}' carregado.")
        self._apply_visibility_rules(); self._update_rho_label(); self._draw_scheme()


def main():
    app = PuncoamentoApp()
    app.mainloop()


if __name__ == "__main__":
    main()
